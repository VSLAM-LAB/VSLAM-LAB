#!/usr/bin/env python3
"""
Module: generate_dataset_xlsx.py
Description: Scans Datasets/dataset_files/*.yaml (and, via get_dataset.py, the
             matching *.py) and writes a formatted .xlsx overview - one row per
             dataset, one Excel Table with a frozen, gradient-filled header
             row (colors sampled from docs/header.png, the VSLAM-LAB logo),
             autofilter and single-line rows - openable directly in Excel or
             Google Sheets. Shares its yaml-scanning core (load_dataset_entries)
             with generate_dataset_table.py's Markdown table; this script adds
             the bibliography columns (Access, Publication, Publication URL,
             Year, BibTeX Key, BibTeX, Authors) sourced from each dataset's
             yaml about: block. Those fields are optional and, for most
             datasets today, blank - see dataset_template.yaml's about: block
             for the field list and fill them in per-dataset as they're
             gathered. Column order/set is defined by _COLUMNS below.
Author: Alejandro Fontan Villacampa
Version: 1.4
Created: 2026-08-04
Updated: 2026-08-06
License: GPLv3
List of Known Bugs: None
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from generate_dataset_table import (
    DEFAULT_DATASET_FILES_DIR,
    DEFAULT_GET_DATASET_PY,
    DEFAULT_README_MD,
    load_dataset_entries,
)

DEFAULT_OUTPUT_XLSX = Path(__file__).resolve().with_name("dataset_table.xlsx")

SHEET_NAME = "Datasets"
TABLE_NAME = "Datasets"

# Sampled from the dominant pixel colors in docs/header.png (the VSLAM-LAB logo/README header),
# left-to-right in the order its "VSLAM-LAB" wordmark actually shades through: mint -> sky blue ->
# periwinkle -> blue-violet -> lavender. Re-sample if the logo is ever redesigned:
#   pixi run -e vslamlab python -c "from PIL import Image; from collections import Counter; \
#     print(Counter(Image.open('docs/header.png').convert('RGB').getdata()).most_common(10))"
_LOGO_GRADIENT = ["B5F3F9", "9BC4FA", "8195FB", "9399ED", "A59DDF"]
_LOGO_BRAND_DARK = "47528A"  # darkened periwinkle, used for text needing contrast on white
_LOGO_TINT = "EEF1FE"  # very light periwinkle tint, used for alternating row banding


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    return "".join(f"{max(0, min(255, c)):02X}" for c in rgb)


def _gradient_hex(t: float) -> str:
    """Linearly interpolate a color at position t in [0, 1] across _LOGO_GRADIENT's stops."""
    t = min(max(t, 0.0), 1.0)
    n = len(_LOGO_GRADIENT) - 1
    seg = t * n
    i = min(int(seg), n - 1)
    local_t = seg - i
    c1, c2 = _hex_to_rgb(_LOGO_GRADIENT[i]), _hex_to_rgb(_LOGO_GRADIENT[i + 1])
    return _rgb_to_hex(tuple(round(c1[k] + (c2[k] - c1[k]) * local_t) for k in range(3)))

# (header, entry key). entry values that are lists are joined with ", " before being written to
# the cell. All columns share one width (_COLUMN_WIDTH below) rather than being sized per-column.
_COLUMNS: list[tuple[str, str]] = [
    ("Dataset", "dataset_name"),
    ("Summary", "display_name"),
    ("Features", "features"),
    ("Modes", "modes"),
    ("Camera Model", "cam_models"),
    ("Web", "homepage"),
    ("Publication", "publication"),
    ("Publication URL", "publication_url"),
    ("Year", "year"),
    ("Authors", "authors"),
    ("BibTeX Key", "bibtex_key"),
    ("BibTeX", "bibtex"),
    ("Maintainer", "maintainer"),
    ("AI-Assisted", "assisted_by"),
    ("Raw Format", "raw_formats"),
    ("Download", "download_labels"),
    ("Download Issues", "issue_ids"),
    ("Access", "access"),
    ("License", "license"),
]
_COLUMN_WIDTH = 20


def _cell_value(entry: dict, key: str) -> str:
    value = entry.get(key, "")
    if isinstance(value, list):
        value = ", ".join(value)
    if isinstance(value, str) and "\n" in value:
        value = " ".join(value.splitlines())  # keep every row a single line (e.g. multi-line bibtex)
    return value


def _write_workbook(entries: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    num_cols = len(_COLUMNS)
    header_font = Font(bold=True, color=_LOGO_BRAND_DARK)
    for col_idx, (header, _key) in enumerate(_COLUMNS, start=1):
        t = (col_idx - 1) / (num_cols - 1) if num_cols > 1 else 0.0
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color=_gradient_hex(t), end_color=_gradient_hex(t), fill_type="solid")
        cell.alignment = Alignment(vertical="center")

    dataset_col = next(i for i, (_h, k) in enumerate(_COLUMNS, start=1) if k == "dataset_name")
    tint_fill = PatternFill(start_color=_LOGO_TINT, end_color=_LOGO_TINT, fill_type="solid")

    for row_idx, entry in enumerate(entries, start=2):
        row_fill = tint_fill if row_idx % 2 == 0 else None
        for col_idx, (_header, key) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(entry, key))
            cell.alignment = Alignment(vertical="top")
            if row_fill is not None:
                cell.fill = row_fill
            if col_idx == dataset_col:
                cell.font = Font(bold=True, color=_LOGO_BRAND_DARK)

        homepage = entry.get("homepage", "")
        if homepage:
            web_col = next(i for i, (_h, k) in enumerate(_COLUMNS, start=1) if k == "homepage")
            web_cell = ws.cell(row=row_idx, column=web_col)
            web_cell.hyperlink = homepage
            web_cell.font = Font(color="0000FF", underline="single")

        publication_url = entry.get("publication_url", "")
        if publication_url:
            pub_url_col = next(i for i, (_h, k) in enumerate(_COLUMNS, start=1) if k == "publication_url")
            pub_url_cell = ws.cell(row=row_idx, column=pub_url_col)
            pub_url_cell.hyperlink = publication_url
            pub_url_cell.font = Font(color="0000FF", underline="single")

    for col_idx in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTH

    ws.freeze_panes = "A2"

    last_row = len(entries) + 1
    last_col_letter = get_column_letter(num_cols)
    table = Table(displayName=TABLE_NAME, ref=f"A1:{last_col_letter}{last_row}")
    # showRowStripes=False: banding is hand-painted above (_LOGO_TINT) instead of a built-in
    # table style, so its color matches the logo instead of Excel's default blue/gray presets.
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False, showFirstColumn=False, showLastColumn=False
    )
    ws.add_table(table)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset-files-dir", type=Path, default=DEFAULT_DATASET_FILES_DIR,
                         help="Directory containing dataset_*.yaml/.py files")
    parser.add_argument("--get-dataset-py", type=Path, default=DEFAULT_GET_DATASET_PY,
                         help="Path to get_dataset.py (used to map dataset name -> implementing class/file)")
    parser.add_argument("--readme", type=Path, default=DEFAULT_README_MD,
                         help="Path to README.md (used to look up each dataset's Features tags)")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_XLSX,
                         help="Output .xlsx file path")
    args = parser.parse_args()

    entries = load_dataset_entries(args.dataset_files_dir, args.get_dataset_py, args.readme)
    _write_workbook(entries, args.output)

    print(f"Wrote {len(entries)} dataset rows to {args.output}")


if __name__ == "__main__":
    main()
